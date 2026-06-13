#!/usr/bin/env python3
"""Convert New HSK (2025) Anki-xiehanzi TSV word lists into .xlsx workbooks.

Each source file is a tab-separated list with these columns:
    0 simplified       1 traditional     2 pinyin
    3 zhuyin           4 hsk_level        5 part_of_speech
    6 frequency        7 html_definition_block

The HTML definition block holds one or more readings, each shaped like:
    <div class="char">…</div>
    <span class="pinYinWrapper">…pinyin…</span>
    <ul><li>gloss</li>…</ul>

We flatten that into a readable definition cell, one line per reading:
    pinyin1: gloss; gloss; gloss
    pinyin2: gloss; gloss

Usage:
    python scripts/hsk_txt_to_xlsx.py                 # convert every HSK_Level_*.txt
    python scripts/hsk_txt_to_xlsx.py path/to/file.txt [more.txt ...]

Output: <source_dir>/xlsx/<name>.xlsx (a subfolder beside the source files).
"""
from __future__ import annotations

import glob
import os
import sys

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Default source directory (relative to repo root).
# Repo root is two levels up from this script (HSK-3.0-words-list/Scripts and data/).
REPO_ROOT = os.path.abspath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..")
)
DEFAULT_DIR = os.path.join(
    REPO_ROOT, "HSK-3.0-words-list", "New HSK (2025)", "Anki xiehanzi"
)

HEADERS = [
    "Simplified",
    "Traditional",
    "Pinyin",
    "Zhuyin",
    "HSK Level",
    "Part of Speech",
    "Frequency",
    "Definitions",
]
DEF_COL = 8  # 1-based column index of the Definitions column.
OUT_SUBDIR = "xlsx"  # Workbooks land in this subfolder of the source directory.


def format_definitions(html: str) -> str:
    """Turn the raw HTML definition block into 'pinyin: gloss; gloss' lines."""
    if not html or not html.strip():
        return ""
    soup = BeautifulSoup(html, "html.parser")
    lines: list[str] = []
    for wrapper in soup.find_all("span", class_="pinYinWrapper"):
        # Collapse the nested per-syllable spans into one spaced pinyin string.
        pinyin = " ".join(wrapper.get_text(" ", strip=True).split())
        ul = wrapper.find_next("ul")
        glosses = (
            [li.get_text(" ", strip=True) for li in ul.find_all("li")]
            if ul
            else []
        )
        glosses = [g for g in glosses if g]
        if pinyin and glosses:
            lines.append(f"{pinyin}: {'; '.join(glosses)}")
        elif pinyin:
            lines.append(pinyin)
        elif glosses:
            lines.append("; ".join(glosses))
    if lines:
        # Blank line between readings for readability.
        return "\n\n".join(lines)
    # Fallback: no structured readings, dump any list items we can find.
    glosses = [li.get_text(" ", strip=True) for li in soup.find_all("li")]
    return "; ".join(g for g in glosses if g)


def parse_row(line: str) -> list[str] | None:
    """Split one TSV line into the 8 output columns; None for blank lines."""
    line = line.rstrip("\n").rstrip("\r")
    if not line.strip():
        return None
    parts = line.split("\t")
    # Pad short rows so column access never fails.
    while len(parts) < 8:
        parts.append("")
    return parts[:7] + [format_definitions(parts[7])]


def convert_file(src: str) -> str:
    out_dir = os.path.join(os.path.dirname(src), OUT_SUBDIR)
    os.makedirs(out_dir, exist_ok=True)
    name = os.path.splitext(os.path.basename(src))[0] + ".xlsx"
    dst = os.path.join(out_dir, name)
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(src))[0]

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = 0
    with open(src, encoding="utf-8") as fh:
        for line in fh:
            row = parse_row(line)
            if row is None:
                continue
            ws.append(row)
            rows += 1

    # Wrap the long definition cells and top-align everything.
    def_letter = get_column_letter(DEF_COL)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            wrap = cell.column == DEF_COL
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)

    # Reasonable column widths.
    widths = {
        "A": 12, "B": 12, "C": 14, "D": 16,
        "E": 11, "F": 18, "G": 12, def_letter: 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    wb.save(dst)
    print(f"  {os.path.basename(src)} -> {os.path.basename(dst)}  ({rows} rows)")
    return dst


def main(argv: list[str]) -> int:
    if argv:
        sources = argv
    else:
        sources = sorted(glob.glob(os.path.join(DEFAULT_DIR, "HSK_Level_*.txt")))
    if not sources:
        print("No source .txt files found.", file=sys.stderr)
        return 1
    print(f"Converting {len(sources)} file(s):")
    for src in sources:
        convert_file(src)
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
